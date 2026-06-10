import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data=json.load(open('_research/data.json'))

order={'高':0,'中':1,'低':2}
prefs=['千葉県','茨城県','埼玉県']

wb=Workbook()
# README sheet first
ws0=wb.active; ws0.title='このリストについて'

hdr_fill=PatternFill('solid',fgColor='1F4E78')
hdr_font=Font(bold=True,color='FFFFFF',size=11)
conf_fill={'高':PatternFill('solid',fgColor='C6E0B4'),'中':PatternFill('solid',fgColor='FFF2CC'),'低':PatternFill('solid',fgColor='F8CBAD')}
thin=Side(style='thin',color='BFBFBF')
border=Border(left=thin,right=thin,top=thin,bottom=thin)
cols=[('No',6),('会社名',26),('住所',40),('電話番号',15),('HP',34),('協力会社募集欄',18),('確度',7),('備考（特徴・除外チェック）',60)]

def style_sheet(ws, rows):
    # header
    for c,(name,w) in enumerate(cols,1):
        cell=ws.cell(1,c,name); cell.fill=hdr_fill; cell.font=hdr_font
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border=border
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.row_dimensions[1].height=30
    r=2
    for i,rec in enumerate(rows,1):
        ws.cell(r,1,i)
        ws.cell(r,2,rec['name'])
        ws.cell(r,3,rec['address'])
        ws.cell(r,4,rec['phone'])
        hp=rec['hp']
        hc=ws.cell(r,5,hp)
        if hp.startswith('http'):
            hc.hyperlink=hp; hc.font=Font(color='0563C1',underline='single')
        part=rec.get('partner','不明')
        if part in ('不明','要確認',''): part='要確認（公式HP）'
        ws.cell(r,6,part)
        cf=ws.cell(r,7,rec['conf']); cf.alignment=Alignment(horizontal='center')
        if rec['conf'] in conf_fill: cf.fill=conf_fill[rec['conf']]
        ws.cell(r,8,rec['note'])
        for c in range(1,9):
            cell=ws.cell(r,c); cell.border=border
            cell.alignment=Alignment(vertical='top',wrap_text=True, horizontal='center' if c in (1,7) else 'left')
        r+=1
    ws.freeze_panes='A2'
    ws.auto_filter.ref=f"A1:H{r-1}"

for p in prefs:
    rows=[x for x in data if x['pref']==p]
    rows.sort(key=lambda x:(order.get(x['conf'],3)))
    ws=wb.create_sheet(p)
    style_sheet(ws,rows)

# README content
ws0['A1']='不動産賃貸管理会社リスト（千葉県・茨城県・埼玉県）'
ws0['A1'].font=Font(bold=True,size=14)
notes=[
 '',
 '■ 抽出条件（すべて満たす会社）',
 '  ・不動産賃貸管理がメイン事業（オーナーから賃貸物件の管理を受託）',
 '  ・社内に「管理部」と「リフォーム部（リフォーム/リノベ/工事事業）」の両方を持つ',
 '  ・地域密着型の独立系・地場企業',
 '',
 '■ 除外した会社',
 '  ・建設会社／工務店／ハウスメーカー／リフォーム専業（建築・施工が本業）',
 '  ・大手建設グループ（大和ハウス・積水ハウス系・大東建託・レオパレス・ポラス・新昭和 等）',
 '  ・建設会社/工務店をグループ・併設に持つ不動産会社（例：三共土地建物＝三共矢崎建設、桂不動産＝建築土木、太陽ハウス＝建築部門、GMC＝一級建築士事務所 等）',
 '  ・売買/買取が主で賃貸管理・リフォームが弱い会社',
 '',
 '■ 各シート：千葉県 / 茨城県 / 埼玉県（確度の高い順に並べ替え済み）',
 '',
 '■ 列の説明',
 '  会社名・住所・電話番号・HP・協力会社募集欄・確度・備考',
 '  ・確度：高=賃貸管理+リフォーム+建設系グループ無しを確認 / 中=おおむね該当 / 低=該当だが管理規模やリフォーム部の体制に要確認点あり',
 '',
 '■ ファクトチェック実施済み（2026-06-10）— 主な修正',
 '  ・柏ホーム TEL 04-7144-3661（旧データの番号は誤りを修正）',
 '  ・丸一土地建物 TEL 043-224-6561 / 松堀不動産 住所=東松山市箭弓町2-3-2・TEL 0493-24-1111',
 '  ・大みか不動産 TEL 0294-53-5878 / 一誠商事 TEL 029-852-6611 / ワンステップハウス 大宮区桜木町1-1-11・048-650-3100',
 '  ・吉田不動産＝本社は千葉県市川市南行徳(東西線地盤)と判明し【埼玉→千葉】へ訂正。大宮の同名社は別法人(土地分譲)で対象外',
 '  ・GMC（結城）＝実体は一級建築士事務所(建築設計・施工)のため除外',
 '',
 '■ 協力会社募集欄について（重要）',
 '  ・各社名で「協力会社/協力業者/職人募集」を個別に検索確認（主要〜中堅31社）。',
 '  ・公開の募集ページが確認できたのは【サカイ・エージェンシー(熊谷)】1社のみ → https://sakai-ag.com/partner/',
 '  ・他社は公開募集ページが無く「無」。＝Webフォームではなく直接打診(電話/訪問)が必要、という結論。',
 '  ・吉岡地所(松戸)はリフォームを外部の協力会社へ発注する運用のため、直接営業の好機。',
 '',
 '■ 注意',
 '  ・電話番号「要確認」は公式番号を特定できなかった社（公式HPに掲載あり）。',
 '  ・社数：千葉18 / 茨城16 / 埼玉14（計48社／条件を満たす地場企業を優先。埼玉の増補・追加調査も可能）',
 '  ・調査日：2026-06-10（検索ベース。最終判断は各社公式HP/直接確認を推奨）',
]
for i,t in enumerate(notes,3):
    c=ws0.cell(i,1,t)
    if t.startswith('■'): c.font=Font(bold=True,size=11,color='1F4E78')
ws0.column_dimensions['A'].width=110

fn='不動産賃貸管理会社リスト_千葉_茨城_埼玉_2026.xlsx'
wb.save(fn)
print('saved',fn)
import os; print('bytes',os.path.getsize(fn))
